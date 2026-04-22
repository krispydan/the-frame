#!/usr/bin/env python3
"""
Sync FOB unit costs from a factory PO spreadsheet into
catalog_skus.cost_price on prod.

These costs feed Shopify's "Cost per item" column (via a landed-cost
multiplier in src/modules/catalog/lib/export/shopify.ts), which in
turn drives Shopify's margin/Finance analytics. If the DB costs drift
from reality every margin number in Shopify becomes fiction.

See docs/fob-cost-sync.md for the full runbook including:
  - Where the spreadsheet comes from
  - What format the spreadsheet needs to be in
  - How the landed-cost multiplier relates to these FOB values
  - How to verify the update worked

Usage:
  # Default: reads the canonical consolidated PO spreadsheet
  python3 scripts/sync-fob-costs.py

  # Or point at a specific file (an .xlsx with a "PO Master" sheet)
  python3 scripts/sync-fob-costs.py --path ~/Downloads/Factory_PO_Consolidated.xlsx

  # Dry run: parse + diff only, don't write
  python3 scripts/sync-fob-costs.py --dry

Spreadsheet contract (PO Master sheet):
  - Header row on row 1.
  - Column C (`JAXY SKU`) holds the SKU code, e.g. "JX1001-BLK".
  - Column J (`Unit Cost`) holds the FOB unit price in USD.
  - Rows with empty SKU or empty cost are skipped.

The script:
  1. Parses the spreadsheet into a {sku: fob_cost} map.
  2. Reads current catalog_skus from prod via the system.query MCP
     tool (read-only admin path).
  3. Diffs old vs new costs. Prints a summary.
  4. Writes changes via a single CASE WHEN UPDATE through the same
     MCP tool (writes ARE allowed for admin-keyed sessions).
  5. Re-reads and reports min/max/avg of the new values.

Auth: uses the Claude Code MCP API key already baked into
scripts/sync-images-to-prod.py. Same key, same endpoint.
"""
import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

API_BASE = "https://theframe.getjaxy.com"
MCP_KEY = "d7b54adb537de1f8a7fe41cffc3b595f39b37490a47edadfe8661e04a55cebc6"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/124.0 Safari/537.36"

# Canonical location — exported from the Factory PO tool.
DEFAULT_XLSX = os.path.expanduser(
    "~/Library/Application Support/Claude/local-agent-mode-sessions/"
    "145b7476-f042-4559-bda1-f5df13fe33c1/"
    "238b3d6f-b7b8-44ac-94c5-65a59e71b668/"
    "local_0e8f502f-4a82-4f9f-a00d-a15ff1cdc28d/outputs/"
    "Factory_PO_Consolidated.xlsx"
)


def mcp_query(sql: str) -> str:
    """Run SQL through the prod MCP system.query tool. Returns the raw
    JSON-encoded result content string. Works for both SELECT and
    UPDATE statements."""
    req = urllib.request.Request(
        f"{API_BASE}/api/mcp",
        data=json.dumps({
            "jsonrpc": "2.0", "id": 1, "method": "tools/call",
            "params": {"name": "system.query", "arguments": {"sql": sql}},
        }).encode(),
        headers={
            "x-api-key": MCP_KEY,
            "Content-Type": "application/json",
            "User-Agent": UA,
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        body = json.loads(resp.read())
    return body.get("result", {}).get("content", [{}])[0].get("text", "")


def parse_spreadsheet(path: Path) -> dict[str, float]:
    """Return {sku: unit_cost_usd} from the PO Master sheet."""
    if not path.exists():
        print(f"ERROR: spreadsheet not found at {path}", file=sys.stderr)
        sys.exit(2)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "PO Master" not in wb.sheetnames:
        print(f"ERROR: expected a 'PO Master' sheet, got {wb.sheetnames}", file=sys.stderr)
        sys.exit(2)

    ws = wb["PO Master"]
    out: dict[str, float] = {}
    # Row 1 = header. Col C = JAXY SKU, Col J = Unit Cost.
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 10:
            continue
        sku, cost = r[2], r[9]
        if not sku or cost is None:
            continue
        out[str(sku).strip()] = float(cost)
    return out


def fetch_current_costs() -> dict[str, float | None]:
    raw = mcp_query("SELECT sku, cost_price FROM catalog_skus ORDER BY sku")
    return {r["sku"]: r["cost_price"] for r in json.loads(raw)}


def build_update_sql(pairs: dict[str, float]) -> str:
    cases: list[str] = []
    skus: list[str] = []
    for sku, cost in pairs.items():
        safe_sku = sku.replace("'", "''")
        cases.append(f"WHEN '{safe_sku}' THEN {cost}")
        skus.append(f"'{safe_sku}'")
    return (
        "UPDATE catalog_skus SET cost_price = CASE sku\n    "
        + "\n    ".join(cases)
        + "\n    ELSE cost_price END "
        + f"WHERE sku IN ({','.join(skus)})"
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--path", default=DEFAULT_XLSX,
                    help="Path to the consolidated PO .xlsx (default: canonical location)")
    ap.add_argument("--dry", action="store_true", help="Print diff but don't write")
    args = ap.parse_args()

    path = Path(args.path).expanduser()
    print(f"Reading: {path}")
    pairs = parse_spreadsheet(path)
    print(f"Parsed  {len(pairs)} SKU → FOB rows "
          f"(min=${min(pairs.values()):.2f}, max=${max(pairs.values()):.2f}, "
          f"avg=${sum(pairs.values()) / len(pairs):.2f})")

    print("Reading current DB costs via MCP...")
    current = fetch_current_costs()
    db_skus = set(current)
    sheet_skus = set(pairs)

    only_db = db_skus - sheet_skus
    only_sheet = sheet_skus - db_skus
    if only_db:
        print(f"  WARN {len(only_db)} SKUs in DB but NOT in spreadsheet (will be left unchanged): {sorted(only_db)[:10]}")
    if only_sheet:
        print(f"  WARN {len(only_sheet)} SKUs in spreadsheet but NOT in DB (will be skipped): {sorted(only_sheet)[:10]}")

    overlap = sheet_skus & db_skus
    changes = {sku: pairs[sku] for sku in overlap if (current.get(sku) or 0) != pairs[sku]}
    unchanged = len(overlap) - len(changes)

    print(f"\nDiff: {len(changes)} updates, {unchanged} unchanged")
    sample = list(changes.items())[:10]
    for sku, new in sample:
        old = current.get(sku)
        old_s = "NULL" if old is None else f"${old:.2f}"
        print(f"  {sku:14} {old_s:>7} → ${new:.2f}")

    if not changes:
        print("Nothing to update. Done.")
        return

    if args.dry:
        print("\n--dry: skipping write. Re-run without --dry to apply.")
        return

    sql = build_update_sql({s: c for s, c in pairs.items() if s in overlap})
    print(f"\nWriting {len(overlap)} rows via MCP...")
    result = mcp_query(sql)
    print(f"  {result}")

    print("\nVerification:")
    verify = json.loads(mcp_query(
        "SELECT MIN(cost_price) mn, MAX(cost_price) mx, "
        "ROUND(AVG(cost_price), 4) av FROM catalog_skus"
    ))[0]
    print(f"  min=${verify['mn']:.2f} max=${verify['mx']:.2f} avg=${verify['av']:.2f}")

    # Spot-check a handful of SKUs
    spot = sorted(changes)[:5]
    quoted = ",".join("'" + s + "'" for s in spot)
    sample_sql = (
        "SELECT sku, cost_price, ROUND(cost_price * 1.25 + 0.50, 2) AS landed "
        + f"FROM catalog_skus WHERE sku IN ({quoted}) ORDER BY sku"
    )
    spot_rows = json.loads(mcp_query(sample_sql))
    print("\n  SKU            FOB    Landed (Shopify Cost per item)")
    for r in spot_rows:
        print(f"  {r['sku']:14} ${r['cost_price']:.2f}   ${r['landed']:.2f}")

    print("\nDone. Remind user to re-export + re-upload Shopify CSV.")


if __name__ == "__main__":
    main()
