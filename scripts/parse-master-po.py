#!/usr/bin/env python3
"""
Parse Jaxy Master PO Excel into a JSON payload for
POST /api/admin/catalog/import-po.

Scope (per Daniel 2026-06-17):
  - Only NEW products are emitted (Status=New). Reorders are skipped
    so existing prod SKUs (under the old `JX3003-BRW` naming) are not
    duplicated by new `JX3003-S-BRW` rows.
  - Sunglasses: one colorway SKU per row.
  - Reading glasses: one colorway SKU per row + the 7 power-variant
    SKUs (1 blue-light coating + 6 diopter powers) from the Reading
    Glasses sheet.
  - 4-pack data is omitted (warehouse-only, no catalog schema for it
    yet).

Color code → display name mapping (best-effort; rename in UI if any
look off):
  See COLOR_NAMES below.

Usage:
  ./scripts/parse-master-po.py \\
    --in  /Users/danielseeff/Downloads/Jaxy-Master-PO-Data.xlsx \\
    --out /Users/danielseeff/GitHub/the-frame/scripts/po-import-payload.json
"""
import argparse
import json
import sys
from collections import defaultdict
import openpyxl

COLOR_NAMES = {
    "BLK": "Black", "TOR": "Tortoise", "BRW": "Brown", "BLU": "Blue",
    "GRN": "Green", "OLV": "Olive", "PNK": "Pink", "GRY": "Grey",
    "RST": "Rust", "RED": "Red", "SND": "Sand", "BUR": "Burgundy",
    "AMB": "Amber", "TEA": "Teal", "TEL": "Teal", "GLD": "Gold",
    "CRY": "Crystal", "SLV": "Silver", "PUR": "Purple", "ORG": "Orange",
    "COR": "Coral",
    # Best-guess; rename in UI if wrong:
    "TGN": "Tort/Green", "ATO": "Amber/Tortoise",
    "CHA": "Champagne", "LTO": "Light Tortoise",
}

# Reading-glasses power column values → (readingPower:float, hasBlueLight:bool)
def parse_power(label: str):
    s = (label or "").strip()
    if "blue light" in s.lower():
        # "0.00 Blue Light" — no magnification, just blue-light coating
        return (0.0, True)
    # "+1.00", "+1.50", etc.
    if s.startswith("+"):
        return (float(s[1:]), False)
    return None


def color_name(code: str) -> str:
    return COLOR_NAMES.get(code, code)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="infile", required=True)
    ap.add_argument("--out", dest="outfile", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.infile, data_only=True)
    products_ws = wb["All Products"]
    readers_ws = wb["Reading Glasses - All Powers"]

    # 1) Index reading-glasses power rows by base SKU + color
    #    so we can attach them to colorways below.
    power_rows = defaultdict(list)  # (baseSku, colorCode) → list[{power, hasBL, sku, upc, cost}]
    for r in readers_ws.iter_rows(values_only=True, min_row=5):
        if not r[0] or not r[5]:
            continue
        factory, prod_name, base_sku, color, power_label, power_sku, upc, cost, _ = r[:9]
        parsed = parse_power(power_label)
        if parsed is None:
            print(f"[warn] skipping unparseable power: {power_label!r} on {power_sku}", file=sys.stderr)
            continue
        rp, has_bl = parsed
        power_rows[(base_sku, color)].append({
            "sku": power_sku,
            "upc": str(upc) if upc else None,
            "costPrice": float(cost) if cost is not None else None,
            "readingPower": rp,
            "hasBlueLightFilter": has_bl,
        })

    # 2) Walk All Products sheet, building one entry per (skuPrefix, name).
    #    Group colorways under each product.
    products = {}  # skuPrefix → product entry
    skipped_reorder = 0
    for r in products_ws.iter_rows(values_only=True, min_row=5):
        if not r[0] or not r[3]:
            continue
        factory, po, name, sku, color_code, ptype, upc, cost, qty, ext, status = r[:11]
        if status != "New":
            skipped_reorder += 1
            continue

        parts = sku.split("-")
        prefix = parts[0]            # JX1019
        type_letter = parts[1] if len(parts) > 1 else None  # S or R

        if prefix not in products:
            products[prefix] = {
                "skuPrefix": prefix,
                "name": name,
                "type": "reading_glasses" if ptype == "Reading Glasses" else "sunglasses",
                "factory": factory,
                "po": po,
                "skus": [],
            }
        product = products[prefix]

        colorway = {
            "sku": sku,
            "colorCode": color_code,
            "colorName": color_name(color_code),
            "upc": str(upc) if upc else None,
            "costPrice": float(cost) if cost is not None else None,
            "qty": int(qty) if qty is not None else None,
        }
        # For reading glasses, attach the 7 power variants
        if product["type"] == "reading_glasses":
            # Reading rows use a different SKU shape: base_sku is e.g.
            # JX1019-R-BLK (without power suffix) — that's how power_rows
            # are keyed. Our All Products row's `sku` is the same form.
            colorway["powerVariants"] = power_rows.get((sku, color_code), [])
            if not colorway["powerVariants"]:
                print(f"[warn] no power variants for {sku} / {color_code}", file=sys.stderr)
        product["skus"].append(colorway)

    payload = {
        "generatedFrom": args.infile,
        "generatedAt": "2026-06-17",
        "skippedReorderRows": skipped_reorder,
        "products": list(products.values()),
    }

    with open(args.outfile, "w") as f:
        json.dump(payload, f, indent=2)

    n_products = len(payload["products"])
    n_colorways = sum(len(p["skus"]) for p in payload["products"])
    n_powers = sum(
        len(c.get("powerVariants", []))
        for p in payload["products"] for c in p["skus"]
    )
    print(f"✓ Wrote {args.outfile}")
    print(f"  Products:     {n_products}")
    print(f"  Colorways:    {n_colorways}")
    print(f"  Power SKUs:   {n_powers}")
    print(f"  Reorder rows skipped: {skipped_reorder}")


if __name__ == "__main__":
    main()
